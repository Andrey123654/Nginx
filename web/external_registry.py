"""Safe import of Edge/DNAT publication registries from XLSX."""

import io
import ipaddress
import re
import zipfile

from openpyxl import load_workbook


HEADER_ALIASES = {
    "rule_id": {"id правила", "ид правила", "rule id", "id"},
    "rule_type": {"тип", "type", "action", "действие"},
    "scope": {
        "источник / интерфейс", "источник/интерфейс", "интерфейс",
        "applied on", "применено на", "область применения",
    },
    "external_ip": {"original ip address", "original ip", "исходный ip", "внешний ip"},
    "external_port": {"внешний порт", "external port", "original port", "исходный порт"},
    "internal_ip": {
        "внутренний ip", "internal ip", "translated ip address", "translated ip",
        "транслированный ip",
    },
    "internal_port": {
        "внутренний порт", "internal port", "translated port", "транслированный порт",
    },
    "protocol": {"протокол", "protocol"},
    "enabled": {"enabled", "включено", "активно"},
    "visibility": {"visibility", "zone", "зона", "область видимости"},
}
REQUIRED_FIELDS = {
    "rule_id", "rule_type", "scope", "external_port", "internal_ip", "internal_port", "protocol",
}
MAX_REGISTRY_ROWS = 10_000
MAX_ZIP_MEMBERS = 500
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024

EXTERNAL_SCOPE_MARKERS = (
    "internet", "external", "outside", "public", "wan", "untrust", "интернет", "внешн",
)
INTERNAL_SCOPE_MARKERS = (
    "internal", "inside", "intranet", "private", "trusted", "trust", "lan", "local",
    "corp", "office", "внутрен", "локальн",
)
GENERIC_SCOPE_VALUES = {"", "any", "all", "current", "global", "любой", "все", "текущий"}


def _normalized(value):
    return " ".join(str(value or "").strip().lower().replace("ё", "е").split())


def _field_map(row):
    values = {_normalized(value): index for index, value in enumerate(row) if _normalized(value)}
    result = {}
    for field, aliases in HEADER_ALIASES.items():
        index = next((values[alias] for alias in aliases if alias in values), None)
        if index is not None:
            result[field] = index
    return result if REQUIRED_FIELDS <= result.keys() else None


def _cell(row, mapping, field):
    index = mapping.get(field)
    return row[index] if index is not None and index < len(row) else None


def _identifier(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip()


def _port(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if int(value) == value and 1 <= int(value) <= 65535:
            return int(value), None
        raise ValueError(f"порт вне диапазона: {value}")
    text = str(value or "").strip()
    if _normalized(text) in {"any", "любой", "все", "*"}:
        return "any", None
    try:
        address = str(ipaddress.ip_address(text))
        return "any", address
    except ValueError:
        pass
    if text.isdigit() and 1 <= int(text) <= 65535:
        return int(text), None
    match = re.fullmatch(r"(\d{1,5})\s*[-–—:]\s*(\d{1,5})", text)
    if match:
        first, last = int(match.group(1)), int(match.group(2))
        if 1 <= first <= last <= 65535:
            return f"{first}-{last}", None
    raise ValueError(f"некорректный порт: {text or 'пусто'}")


def _ip(value, row_number, label):
    text = str(value or "").strip()
    try:
        return str(ipaddress.ip_address(text))
    except ValueError as exc:
        raise ValueError(f"строка {row_number}: некорректный {label}") from exc


def _network_from_text(value):
    text = str(value or "").strip()
    match = re.search(r"(?<![\d.])(\d{1,3}(?:\.\d{1,3}){3})(?:[/-](\d{1,2}))?(?![\d.])", text)
    if not match:
        return None
    candidate = match.group(1) + ("/" + match.group(2) if match.group(2) else "")
    try:
        return str(ipaddress.ip_network(candidate, strict=False)) if "/" in candidate else str(ipaddress.ip_address(candidate))
    except ValueError:
        return None


def _enabled(value):
    if value in (None, ""):
        return True
    return _normalized(value) not in {"no", "false", "0", "off", "нет", "выключено", "disabled"}


def _visibility(scope, explicit=None):
    requested = _normalized(explicit)
    if requested in {"external", "внешняя", "наружу"}:
        return "external", "зона указана в выгрузке", "high"
    if requested in {"internal", "внутренняя", "внутрь"}:
        return "internal", "зона указана в выгрузке", "high"
    value = _normalized(scope)
    if any(marker in value for marker in EXTERNAL_SCOPE_MARKERS):
        return "external", f"область применения «{scope}» распознана как Internet/Outside", "high"
    if any(marker in value for marker in INTERNAL_SCOPE_MARKERS):
        return "internal", f"область применения «{scope}» распознана как Inside/LAN", "high"
    if value in GENERIC_SCOPE_VALUES:
        return "unknown", "область применения не позволяет определить периметр", "low"
    return "internal", f"правило ограничено именованной областью «{scope}», не являющейся Internet/Outside", "medium"


def _validate_archive(raw):
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ZIP_MEMBERS:
                raise ValueError("слишком много объектов внутри XLSX")
            if sum(item.file_size for item in members) > MAX_UNCOMPRESSED_BYTES:
                raise ValueError("XLSX превышает безопасный распакованный размер")
            names = {item.filename.lower() for item in members}
            if any(name.endswith("vbaproject.bin") for name in names):
                raise ValueError("XLSX с макросами не поддерживается")
            if archive.testzip() is not None:
                raise ValueError("XLSX повреждён")
    except zipfile.BadZipFile as exc:
        raise ValueError("файл не является корректным XLSX") from exc


def parse_external_registry_xlsx(raw, filename="edge-dnat-rules.xlsx"):
    """Parse both the legacy registry and the current Edge export."""
    _validate_archive(raw)
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise ValueError("не удалось прочитать книгу XLSX") from exc
    try:
        selected = None
        for sheet in workbook.worksheets:
            for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), 1):
                mapping = _field_map(row)
                if mapping:
                    selected = (sheet, row_number, mapping)
                    break
            if selected:
                break
        if not selected:
            raise ValueError(
                "не найдена строка заголовков Edge/DNAT; ожидаются ID, Action/Тип, Applied on/Интерфейс, "
                "Original/Внешний Port, Translated/Внутренний IP и Port, Protocol"
            )

        sheet, header_row, mapping = selected
        grouped = {}
        source_rows = 0
        disabled_ignored = 0
        non_dnat_ignored = 0
        exact_duplicates = 0
        seen_rows = set()
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not any(value not in (None, "") for value in row):
                continue
            if source_rows >= MAX_REGISTRY_ROWS:
                raise ValueError(f"реестр содержит более {MAX_REGISTRY_ROWS} строк")
            source_rows += 1
            if not _enabled(_cell(row, mapping, "enabled")):
                disabled_ignored += 1
                continue
            rule_type = str(_cell(row, mapping, "rule_type") or "").strip().upper()
            if rule_type not in {"DNAT", "ДНАТ"}:
                non_dnat_ignored += 1
                continue
            rule_id = _identifier(_cell(row, mapping, "rule_id"))
            if not rule_id:
                raise ValueError(f"строка {row_number}: не указан ID правила")
            internal_ip = _ip(_cell(row, mapping, "internal_ip"), row_number, "Translated/Internal IP")
            external_port, address_from_port = _port(_cell(row, mapping, "external_port"))
            internal_port, _ = _port(_cell(row, mapping, "internal_port"))
            protocol = _normalized(_cell(row, mapping, "protocol"))
            protocol = {"6": "tcp", "17": "udp", "любой": "any", "все": "any", "*": "any"}.get(protocol, protocol)
            if protocol not in {"tcp", "udp", "any"}:
                raise ValueError(f"строка {row_number}: протокол должен быть tcp, udp или Any")
            scope = str(_cell(row, mapping, "scope") or "").strip()
            external_ip = None
            if "external_ip" in mapping:
                external_ip = _ip(_cell(row, mapping, "external_ip"), row_number, "Original/External IP")
            external_ip = external_ip or address_from_port or _network_from_text(scope)
            zone, zone_basis, zone_confidence = _visibility(scope, _cell(row, mapping, "visibility"))
            row_signature = (rule_id, rule_type, scope, external_ip, external_port, internal_ip, internal_port, protocol, zone)
            if row_signature in seen_rows:
                exact_duplicates += 1
                continue
            seen_rows.add(row_signature)
            signature = (rule_type, external_ip, external_port, internal_ip, internal_port, protocol)
            item = grouped.setdefault(signature, {
                "rule_id": rule_id, "rule_ids": [], "rule_type": rule_type,
                "source_interface": scope, "scope_names": [], "external_network": external_ip,
                "external_port": external_port, "internal_ip": internal_ip,
                "internal_port": internal_port, "protocol": protocol,
                "visibility": [], "visibility_evidence": [], "sheet": sheet.title, "source_rows": [],
            })
            item["rule_ids"].append(rule_id)
            if scope and scope not in item["scope_names"]:
                item["scope_names"].append(scope)
            if zone not in item["visibility"]:
                item["visibility"].append(zone)
            item["visibility_evidence"].append({
                "scope": scope, "zone": zone, "basis": zone_basis, "confidence": zone_confidence,
            })
            item["source_rows"].append(row_number)

        rules = list(grouped.values())
        if not rules:
            raise ValueError("реестр не содержит включённых правил DNAT")
        for item in rules:
            item["rule_ids"].sort()
            item["visibility"].sort()
            item["source_rows"].sort()
        zone_counts = {zone: 0 for zone in ("external", "internal", "both", "unknown")}
        for item in rules:
            known = set(item["visibility"]) - {"unknown"}
            key = "both" if known == {"external", "internal"} else next(iter(known), "unknown")
            zone_counts[key] += 1
        return {
            "kind": "edge-dnat-registry", "format": "edge" if "external_ip" in mapping else "legacy",
            "source": filename, "sheet": sheet.title, "header_row": header_row,
            "source_rows": source_rows, "rules": rules,
            "collapsed_rules": max(0, source_rows - disabled_ignored - non_dnat_ignored - len(rules)),
            "duplicates_ignored": exact_duplicates, "disabled_ignored": disabled_ignored,
            "non_dnat_ignored": non_dnat_ignored, "zone_counts": zone_counts,
        }
    finally:
        workbook.close()
